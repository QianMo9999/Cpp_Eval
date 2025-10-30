"""
评价结果保存模块
支持保存为文本、Markdown、PDF和Excel格式
"""
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import json
import re


class ResultSaver:
    """评价结果保存器"""

    def __init__(self, output_dir: str = "./output"):
        """
        初始化保存器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def save_individual_report(
        self,
        student_name: str,
        file_name: str,
        evaluation: str,
        week: str = "02",
        format: str = "markdown"
    ) -> str:
        """
        保存单个学生的评价报告

        Args:
            student_name: 学生姓名
            file_name: 代码文件名
            evaluation: 评价内容
            week: 周次
            format: 保存格式 (markdown, txt)

        Returns:
            保存的文件路径
        """
        # 创建周次目录
        week_dir = os.path.join(self.output_dir, f"第{week}周")
        os.makedirs(week_dir, exist_ok=True)

        # 确定文件扩展名
        ext = ".md" if format == "markdown" else ".txt"

        # 生成文件名
        safe_student_name = student_name.replace('/', '_').replace('\\', '_')
        file_path = os.path.join(week_dir, f"{safe_student_name}_评价报告{ext}")

        # 准备内容
        if format == "markdown":
            content = self._format_markdown_report(student_name, file_name, evaluation, week)
        else:
            content = self._format_text_report(student_name, file_name, evaluation, week)

        # 保存文件
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)

        print(f"✓ 已保存评价报告: {file_path}")
        return file_path

    def _format_markdown_report(
        self,
        student_name: str,
        file_name: str,
        evaluation: str,
        week: str
    ) -> str:
        """格式化Markdown报告"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        return f"""# C++作业评价报告

---

**学生姓名**: {student_name}
**作业周次**: 第{week}周
**代码文件**: {file_name}
**评价时间**: {timestamp}

---

{evaluation}

---
"""

    def _format_text_report(
        self,
        student_name: str,
        file_name: str,
        evaluation: str,
        week: str
    ) -> str:
        """格式化文本报告"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        return f"""====================================
C++作业评价报告
====================================

学生姓名: {student_name}
作业周次: 第{week}周
代码文件: {file_name}
评价时间: {timestamp}

------------------------------------

{evaluation}

------------------------------------

"""

    def save_summary_excel(
        self,
        results: List[Dict],
        week: str = "02",
        filename: str = None
    ) -> str:
        """
        保存汇总Excel表格

        Args:
            results: 评价结果列表
                [
                    {
                        'student_name': '张三',
                        'file_name': 'code.cpp',
                        'evaluation': '评价内容...',
                        'score': 85,  # 可选
                        'timestamp': '2024-01-01 10:00:00'
                    },
                    ...
                ]
            week: 周次
            filename: 自定义文件名（可选）

        Returns:
            保存的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise Exception("请安装openpyxl库: pip install openpyxl")

        # 生成文件名
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"第{week}周_评价汇总_{timestamp}.xlsx"

        file_path = os.path.join(self.output_dir, filename)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"第{week}周评价汇总"

        # 设置表头
        headers = ['序号', '学号', '学生姓名', '文件名', '评分', '状态', '评价时间', '评价内容']
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充数据
        for idx, result in enumerate(results, start=1):
            # 确定状态
            status = result.get('status', 'evaluated')
            status_text = {
                'evaluated': '已评价',
                'not_submitted': '未提交',
                'failed': '评价失败'
            }.get(status, '未知')

            ws.append([
                idx,
                result.get('student_id', ''),
                result.get('student_name', '未知'),
                result.get('file_name', ''),
                result.get('score', ''),
                status_text,
                result.get('timestamp', ''),
                result.get('evaluation', '')
            ])

        # 调整列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 18  # 学号
        ws.column_dimensions['C'].width = 12  # 学生姓名
        ws.column_dimensions['D'].width = 20  # 文件名
        ws.column_dimensions['E'].width = 10  # 评分
        ws.column_dimensions['F'].width = 12  # 状态
        ws.column_dimensions['G'].width = 20  # 评价时间
        ws.column_dimensions['H'].width = 60  # 评价内容

        # 设置文本对齐
        for row in ws.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal='center')  # 序号
            row[1].alignment = Alignment(horizontal='center')  # 学号
            row[2].alignment = Alignment(horizontal='center')  # 学生姓名
            row[4].alignment = Alignment(horizontal='center')  # 评分
            row[5].alignment = Alignment(horizontal='center')  # 状态
            row[6].alignment = Alignment(horizontal='center')  # 时间
            row[7].alignment = Alignment(wrap_text=True, vertical='top')  # 评价内容

        # 保存文件
        wb.save(file_path)
        print(f"✓ 已保存汇总表格: {file_path}")
        return file_path

    def save_json(self, results: List[Dict], week: str = "02", filename: str = None) -> str:
        """
        保存为JSON格式（便于后续处理）

        Args:
            results: 评价结果列表
            week: 周次
            filename: 自定义文件名（可选）

        Returns:
            保存的文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"第{week}周_评价结果_{timestamp}.json"

        file_path = os.path.join(self.output_dir, filename)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"✓ 已保存JSON结果: {file_path}")
        return file_path

    def save_student_pdf(
        self,
        student_name: str,
        student_id: str,
        evaluations: List[Dict],
        week: str = "02"
    ) -> str:
        """
        为每个学生生成一份PDF报告，包含所有题目的评价
        【新方案】先生成Markdown，然后渲染为HTML，最后转换为PDF

        Args:
            student_name: 学生姓名
            student_id: 学号
            evaluations: 该学生的所有评价列表
                [
                    {
                        'file_name': 'main.cpp',
                        'problem_name': '第1关-求三位数',
                        'code': '代码内容...',
                        'evaluation': '评价内容...',
                        'score': 85,
                        'timestamp': '2024-01-01 10:00:00'
                    },
                    ...
                ]
            week: 周次

        Returns:
            保存的文件路径
        """
        try:
            import markdown
            from weasyprint import HTML, CSS
            from weasyprint.text.fonts import FontConfiguration
            # 尝试导入 Pygments 用于代码高亮
            try:
                from pygments.formatters import HtmlFormatter
                PYGMENTS_AVAILABLE = True
            except ImportError:
                PYGMENTS_AVAILABLE = False
        except ImportError:
            print("⚠ 缺少依赖库，请安装:")
            print("  pip install markdown weasyprint")
            print("  可选（代码高亮）: pip install pygments")
            raise Exception("请安装markdown和weasyprint库")

        # 创建周次目录
        week_dir = os.path.join(self.output_dir, f"第{week}周_PDF")
        os.makedirs(week_dir, exist_ok=True)

        # 生成文件名
        safe_student_name = student_name.replace('/', '_').replace('\\', '_')
        student_display = f"{student_id}_{safe_student_name}" if student_id else safe_student_name

        md_file_path = os.path.join(week_dir, f"{student_display}_评价报告.md")
        pdf_file_path = os.path.join(week_dir, f"{student_display}_评价报告.pdf")

        # 1. 生成Markdown内容
        markdown_content = self._generate_markdown_report(
            student_name, student_id, evaluations, week
        )

        # 保存Markdown文件
        with open(md_file_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)

        # 2. 将Markdown转换为HTML
        md_extensions = [
            'fenced_code',      # 支持```代码块
            'tables',           # 表格支持
            'nl2br',            # 换行支持
            'extra',            # 额外功能
        ]

        # 启用代码高亮
        if PYGMENTS_AVAILABLE:
            md_extensions.append('codehilite')

        md = markdown.Markdown(
            extensions=md_extensions,
            extension_configs={
                'codehilite': {
                    'css_class': 'highlight',
                    'linenums': False,
                    'guess_lang': False,
                }
            }
        )
        html_body = md.convert(markdown_content)

        # 获取 Pygments CSS（如果可用）
        pygments_css = ""
        if PYGMENTS_AVAILABLE:
            # 使用 GitHub 风格的语法高亮
            from pygments.styles import get_style_by_name
            formatter = HtmlFormatter(style='github-dark')
            pygments_css = formatter.get_style_defs('.highlight')

        # 3. 添加CSS样式（完全模拟Markdown预览效果）
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>C++作业评价报告 - {student_name}</title>
    <style>
        @page {{
            size: A4;
            margin: 2cm;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
            line-height: 1.6;
            color: #24292e;
            font-size: 14px;
            max-width: 100%;
        }}

        /* 标题样式 */
        h1 {{
            font-size: 2em;
            font-weight: 600;
            padding-bottom: 0.3em;
            border-bottom: 2px solid #1a5490;
            margin-top: 24px;
            margin-bottom: 16px;
            color: #1a5490;
            text-align: center;
        }}

        h2 {{
            font-size: 1.5em;
            font-weight: 600;
            padding-bottom: 0.3em;
            border-bottom: 1px solid #eaecef;
            margin-top: 24px;
            margin-bottom: 16px;
            color: #2c5282;
        }}

        h3 {{
            font-size: 1.25em;
            font-weight: 600;
            margin-top: 24px;
            margin-bottom: 16px;
            color: #4a5568;
        }}

        /* 表格样式 */
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 16px 0;
            display: table;
            overflow: auto;
        }}

        table th {{
            font-weight: 600;
            padding: 6px 13px;
            border: 1px solid #d0d7de;
            background-color: #f6f8fa;
        }}

        table td {{
            padding: 6px 13px;
            border: 1px solid #d0d7de;
        }}

        table tr {{
            background-color: #ffffff;
            border-top: 1px solid #d0d7de;
        }}

        table tr:nth-child(2n) {{
            background-color: #f6f8fa;
        }}

        /* 代码块样式 */
        pre {{
            background-color: #f6f8fa;
            border-radius: 6px;
            padding: 16px;
            overflow: auto;
            font-size: 13px;
            line-height: 1.45;
            margin: 16px 0;
            border: 1px solid #d0d7de;
        }}

        code {{
            background-color: rgba(175, 184, 193, 0.2);
            padding: 0.2em 0.4em;
            margin: 0;
            font-size: 85%;
            border-radius: 3px;
            font-family: ui-monospace, SFMono-Regular, "SF Mono", Menlo, Consolas, "Liberation Mono", monospace;
            color: #24292e;
        }}

        pre code {{
            background-color: transparent;
            padding: 0;
            font-size: 100%;
            color: #24292e;
            border-radius: 0;
            display: block;
            white-space: pre;
            word-break: normal;
            word-wrap: normal;
            font-family: ui-monospace, SFMono-Regular, "SF Mono", Menlo, Consolas, "Liberation Mono", monospace;
        }}

        /* 列表样式 */
        ul, ol {{
            padding-left: 2em;
            margin-top: 0;
            margin-bottom: 16px;
        }}

        li {{
            margin-top: 0.25em;
        }}

        li + li {{
            margin-top: 0.25em;
        }}

        /* 强调样式 */
        strong {{
            font-weight: 600;
        }}

        em {{
            font-style: italic;
        }}

        /* 链接样式 */
        a {{
            color: #0969da;
            text-decoration: none;
        }}

        a:hover {{
            text-decoration: underline;
        }}

        /* 水平线 */
        hr {{
            height: 0.25em;
            padding: 0;
            margin: 24px 0;
            background-color: #d0d7de;
            border: 0;
        }}

        /* 引用块 */
        blockquote {{
            padding: 0 1em;
            color: #656d76;
            border-left: 0.25em solid #d0d7de;
            margin: 0 0 16px 0;
        }}

        /* 特殊样式 */
        .score {{
            color: #1a7f37;
            font-weight: 600;
        }}

        .problem-section {{
            page-break-inside: avoid;
            margin-bottom: 32px;
        }}

        .footer {{
            margin-top: 48px;
            padding-top: 16px;
            border-top: 1px solid #d0d7de;
            text-align: center;
            color: #656d76;
            font-size: 12px;
        }}

        /* 代码高亮（GitHub风格） */
        .codehilite {{
            background-color: #f6f8fa;
            border-radius: 6px;
            padding: 16px;
            margin: 16px 0;
        }}

        .codehilite pre {{
            background-color: transparent;
            border: none;
            padding: 0;
            margin: 0;
        }}
    </style>
</head>
<body>
{html_body}
</body>
</html>
"""

        # 4. 使用WeasyPrint将HTML转换为PDF
        font_config = FontConfiguration()
        HTML(string=html_content).write_pdf(
            pdf_file_path,
            font_config=font_config
        )

        print(f"✓ 已保存PDF报告: {pdf_file_path}")
        print(f"  (Markdown源文件: {md_file_path})")

        return pdf_file_path

    def _generate_markdown_report(
        self,
        student_name: str,
        student_id: str,
        evaluations: List[Dict],
        week: str
    ) -> str:
        """
        生成Markdown格式的报告内容

        Args:
            student_name: 学生姓名
            student_id: 学号
            evaluations: 评价列表
            week: 周次

        Returns:
            Markdown格式的报告内容
        """
        # 计算总评分
        scores = [e.get('score', 0) for e in evaluations if e.get('score') is not None]
        total_score = sum(scores)
        avg_score = total_score / len(scores) if scores else 0

        # 生成报告头部
        md_content = f"""# C++作业评价报告

| 项目 | 内容 |
|------|------|
| **学生姓名** | {student_name} |
| **学号** | {student_id or '无'} |
| **作业周次** | 第{week}周 |
| **评价时间** | {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} |
| **题目总数** | {len(evaluations)} |
| **平均分** | {avg_score:.1f}/100 |
| **总分** | {total_score}/{len(scores)*100} |

---

"""

        # 添加每道题的评价
        for idx, eval_data in enumerate(evaluations, 1):
            problem_name = eval_data.get('problem_name', eval_data.get('file_name', '未知题目'))
            score = eval_data.get('score', 0)
            student_code = eval_data.get('code', '')
            evaluation_text = eval_data.get('evaluation', '无评价')

            # 题目标题（不加"题目1:"前缀）
            md_content += f"""## {problem_name} (得分: {score}/100)

"""

            # 添加学生代码
            if student_code:
                # 限制代码行数
                code_lines = student_code.split('\n')
                display_code = '\n'.join(code_lines[:50])

                md_content += f"""### 📝 提交代码

```cpp
{display_code}
```

"""
                if len(code_lines) > 50:
                    md_content += f"*（代码共{len(code_lines)}行，仅显示前50行）*\n\n"

            # 清理评价内容
            evaluation_text = evaluation_text.strip()
            # 移除可能的题目标识重复
            evaluation_text = re.sub(r'^【题目\d+:.*?】\s*', '', evaluation_text)
            evaluation_text = re.sub(r'^###\s*题目\d+:.*?\n', '', evaluation_text)

            md_content += evaluation_text + "\n\n"

            # 添加分隔线（除了最后一题）
            if idx < len(evaluations):
                md_content += "---\n\n"

        return md_content

    def _calculate_total_score(self, evaluations: List[Dict]) -> str:
        """
        计算总评分

        Args:
            evaluations: 评价列表

        Returns:
            格式化的分数字符串
        """
        scores = [e.get('score', 0) for e in evaluations if e.get('score') is not None]
        if not scores:
            return "0/0"

        total = sum(scores)
        avg = total / len(scores)
        return f"{avg:.1f} (总分: {total}/{len(scores)*100})"


def main():
    """测试函数"""
    saver = ResultSaver()

    # 测试单个报告
    saver.save_individual_report(
        student_name="张三",
        file_name="test.cpp",
        evaluation="## 总体评分\n- 分数: 85/100\n\n## 优点\n- 代码逻辑清晰\n\n## 需要改进的地方\n- 需要添加注释",
        week="02"
    )

    # 测试汇总表格
    results = [
        {
            'student_name': '张三',
            'file_name': 'test1.cpp',
            'evaluation': '代码优秀',
            'score': 95,
            'timestamp': '2024-01-01 10:00:00'
        },
        {
            'student_name': '李四',
            'file_name': 'test2.cpp',
            'evaluation': '需要改进',
            'score': 75,
            'timestamp': '2024-01-01 10:05:00'
        }
    ]
    saver.save_summary_excel(results, week="02")


if __name__ == "__main__":
    main()